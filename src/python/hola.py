from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook, Workbook
from flask_cors import CORS
import os

app = Flask(__name__)
CORS(app)  # Esto permite solicitudes CORS desde cualquier origen


@app.route('/upload', methods=['POST'])
def upload_file():
    uploaded_file = request.files['file']
    if uploaded_file.filename != '':
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active

        products = []
        map = {
            1: 'sku',
            2: 'title',
            3: 'brand',
            4: 'category',
            5: 'iva',
            6: 'costo'
        }
        row = sheet.max_row
        column = sheet.max_column

        for j in range(2, row + 1):
            prod_info = {}
            for i in range(1, column + 1):
                cell_obj = sheet.cell(row=j, column=i)
                column_key = map.get(i)
                if column_key:
                    prod_info[column_key] = cell_obj.value
            products.append(prod_info)

        return jsonify(products)

    return jsonify({'error': 'No se ha recibido ningún archivo'})

@app.route('/update_file', methods=['POST'])
def update_file():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON data provided"}), 400

        workbook = Workbook()

        for market_name, products in data.items():
            sheet = workbook.create_sheet(title=str(market_name))
            if products:
                id_columnas = list(products[0].keys())
                sheet.append(id_columnas)
                for product in products:
                    row_data = [product[key] for key in id_columnas]
                    sheet.append(row_data)

        default_sheet = workbook.get_sheet_by_name('Sheet')
        workbook.remove(default_sheet)

        excel_file = 'excel_actualizado.xlsx'
        excel_file_path = os.path.join(os.getcwd(), excel_file)
        workbook.save(excel_file_path)

        return send_file(excel_file_path, as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000)
